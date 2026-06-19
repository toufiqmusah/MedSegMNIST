import argparse
import json
import os
import numpy as np
from PIL import Image
from openpyxl import load_workbook
from sklearn.model_selection import KFold
import tifffile

import sys
sys.path.insert(0, os.path.dirname(__file__))
from common import (
    normalize_xray,
    compute_padded_shape,
    pad_to_shape,
    save_npz,
    write_checksum,
)

FLAG = "breast2d"
CLASS_NAME = "BreastSegMNIST2D"
MODALITY = "Mammography"
ANATOMY = "Breast"
DIMENSIONALITY = "2D"

STANDARDISED_SIZES = {
    128: 128,
    256: 256,
    512: 512,
}

LABEL_NAMES = {
    "0": "background",
    "1": "mass",
}

N_FOLDS = 5
SPLIT_SEED = 42


def parse_metadata(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    entries = {}
    for row in ws.iter_rows(min_row=32, values_only=True):
        img_id = row[0]
        if img_id is None:
            continue
        img_id = str(img_id).strip()
        if not img_id.startswith("IMG"):
            continue
        e = {
            "img_id": img_id,
            "view": str(row[1] or "").strip(),
            "tissue": str(row[2] or "").strip(),
            "abnormality": str(row[3] or "").strip(),
            "ab_class": str(row[4] or "").strip(),
        }
        if img_id not in entries:
            entries[img_id] = e
    return entries


def collect_samples(img_dir, mask_dir, metadata):
    all_img_files = sorted(f for f in os.listdir(img_dir) if f.endswith(".tif"))
    mask_ids = set(os.path.splitext(f)[0] for f in os.listdir(mask_dir) if f.endswith(".tif"))

    samples = []
    for fname in all_img_files:
        img_id = os.path.splitext(fname)[0]
        img_path = os.path.join(img_dir, fname)
        has_mask = img_id in mask_ids
        mask_path = os.path.join(mask_dir, fname) if has_mask else None

        meta = {"original_id": img_id}
        if img_id in metadata:
            m = metadata[img_id]
            meta["view"] = m["view"]
            meta["tissue"] = m["tissue"]
            meta["abnormality"] = m["abnormality"]
            meta["ab_class"] = m["ab_class"]
        else:
            meta["view"] = ""
            meta["tissue"] = ""
            meta["abnormality"] = "NORM"
            meta["ab_class"] = "N"

        samples.append({
            "img_path": img_path,
            "mask_path": mask_path,
            "meta": meta,
        })

    return samples


def main(raw_dir, out_dir, sizes):
    organ_dir = os.path.join(out_dir, "breast")
    os.makedirs(organ_dir, exist_ok=True)
    checksum_dir = os.path.join(out_dir, "checksums")
    os.makedirs(checksum_dir, exist_ok=True)

    img_dir = os.path.join(raw_dir, "TIFF Images", "TIFF Images")
    mask_dir = os.path.join(raw_dir, "ROI Masks", "ROI Masks")
    meta_path = os.path.join(raw_dir, "Metadata.xlsx")

    print("Parsing metadata...")
    metadata = parse_metadata(meta_path)
    print(f"  {len(metadata)} unique image entries")

    print("Collecting samples...")
    all_samples = collect_samples(img_dir, mask_dir, metadata)
    n_total = len(all_samples)
    print(f"  Total samples: {n_total}")
    n_with_masks = sum(1 for s in all_samples if s["mask_path"] is not None)
    n_norm = n_total - n_with_masks
    print(f"  With mass masks: {n_with_masks}, Normal (no mask): {n_norm}")

    print("Creating 5-fold CV splits...")
    kf = KFold(n_splits=N_FOLDS, shuffle=True, random_state=SPLIT_SEED)
    cv_folds = {}
    for i, (tr, te) in enumerate(kf.split(np.arange(n_total))):
        cv_folds[f"fold_{i}"] = {"train": tr.tolist(), "test": te.tolist()}

    for i in range(n_total):
        all_samples[i]["meta"]["split"] = "all"

    use_native = "native" in sizes
    fixed_sizes = [s for s in sizes if s != "native"]

    if use_native:
        print("Computing native padded shape...")
        all_shapes = []
        for s in all_samples:
            img = Image.open(s["img_path"])
            all_shapes.append(img.size[::-1])
        padded_shape = compute_padded_shape(all_shapes, percentile=95)
        print(f"  Native padded shape: {padded_shape}")
    else:
        padded_shape = None

    npz_buffers = {}
    for size_val in fixed_sizes:
        ts = STANDARDISED_SIZES[size_val]
        npz_buffers[size_val] = (
            np.zeros((n_total, ts, ts), dtype=np.float32),
            np.zeros((n_total, ts, ts), dtype=np.uint8),
        )
    if use_native:
        native_dir = os.path.join(organ_dir, f"{FLAG}_native")
        os.makedirs(native_dir, exist_ok=True)

    for i, s in enumerate(all_samples):
        pil_img = Image.open(s["img_path"])
        pil_img.load()

        if s["mask_path"] is not None:
            mask_raw = tifffile.imread(s["mask_path"])
            if mask_raw.ndim == 3:
                mask_raw = mask_raw[:, :, 0]
            mask_bin_full = (mask_raw > 128).astype(np.uint8)
        else:
            mask_bin_full = None

        for size_val in fixed_sizes:
            ts = STANDARDISED_SIZES[size_val]
            img_rs_pil = pil_img.resize((ts, ts), Image.LANCZOS)
            img_np = np.array(img_rs_pil, dtype=np.float32)
            if img_np.ndim == 3:
                img_np = img_np[:, :, 0]
            npz_buffers[size_val][0][i] = normalize_xray(img_np)

            if mask_bin_full is not None:
                msk_pil = Image.fromarray(mask_bin_full * 255).resize((ts, ts), Image.NEAREST)
                npz_buffers[size_val][1][i] = (np.array(msk_pil) > 128).astype(np.uint8)
            else:
                npz_buffers[size_val][1][i] = np.zeros((ts, ts), dtype=np.uint8)

        if use_native:
            img_np_full = np.array(pil_img)
            if img_np_full.ndim == 3:
                img_np_full = img_np_full[:, :, 0]
            img_padded = pad_to_shape(img_np_full, padded_shape)
            img_uint8 = np.clip(img_padded, 0, 255).astype(np.uint8)
            msk = mask_bin_full if mask_bin_full is not None else np.zeros(img_np_full.shape[:2], dtype=np.uint8)
            msk = pad_to_shape(msk, padded_shape)
            native_path = os.path.join(native_dir, f"{FLAG}_{i}.npz")
            np.savez_compressed(native_path, image=img_uint8, mask=msk)

        if (i + 1) % 50 == 0:
            print(f"    {i + 1}/{n_total}")

        del pil_img, mask_bin_full

    generated_sizes = []
    for size_val in sizes:
        if size_val == "native":
            generated_sizes.append("native")
            continue
        size_key = str(size_val)
        npz_path = os.path.join(organ_dir, f"{FLAG}_{size_val}.npz")

        images, masks = npz_buffers[size_val]
        print(f"  Saving {size_val} ({images.nbytes / 1e6:.0f} MB)...")
        save_npz(
            npz_path,
            images.astype(np.float32),
            masks.astype(np.uint8),
            np.zeros((0, *images.shape[1:]), dtype=np.float32),
            np.zeros((0, *masks.shape[1:]), dtype=np.uint8),
        )

        write_checksum(npz_path, checksum_dir, f"{FLAG}_{size_key}.sha256")
        generated_sizes.append(size_val)
        del npz_buffers[size_val]

    available_sizes = sorted([s for s in generated_sizes if isinstance(s, int)])
    if "native" in generated_sizes:
        available_sizes += ["native"]

    all_meta = [s["meta"] for s in all_samples]
    meta = {
        "flag": FLAG,
        "class_name": CLASS_NAME,
        "name": "DMID (Digital Mammography Dataset for Breast Cancer Diagnosis Research)",
        "version": "1.0.0",
        "dimensionality": DIMENSIONALITY,
        "modality": MODALITY,
        "anatomy": ANATOMY,
        "source_urls": [
            "https://www.kaggle.com/datasets/orvile/dmid-breast-cancer-mammography-dataset",
            "https://figshare.com/articles/dataset/_b_Digital_mammography_Dataset_for_Breast_Cancer_Diagnosis_Research_DMID_b_DMID_rar/24522883",
        ],
        "license": "CC-BY-SA-4.0",
        "redistribution_allowed": True,
        "paper_doi": "https://doi.org/10.1007/s13534-023-00339-y",
        "split_seed": SPLIT_SEED,
        "split_strategy": "5-fold_cv",
        "source_counts": {"tiff_images": 511, "roi_masks": 269},
        "available_sizes": available_sizes,
        "native_padded_shape": list(padded_shape) if padded_shape else None,
        "native_percentile_box": 95,
        "standardised_sizes": {
            str(k): {"shape": [v, v]}
            for k, v in STANDARDISED_SIZES.items()
            if k in [s for s in generated_sizes if isinstance(s, int)]
        },
        "normalization": "percentile_clip_zscore",
        "normalization_percentiles": [1.0, 99.0],
        "label_names": LABEL_NAMES,
        "n_total": n_total,
        "n_folds": N_FOLDS,
        "cv_folds": cv_folds,
        "per_sample_metadata": all_meta,
    }

    with open(os.path.join(organ_dir, f"{FLAG}.json"), "w") as f:
        json.dump(meta, f, indent=2)


def parse_size(s):
    if s == "native":
        return "native"
    return int(s)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=f"Preprocess DMID mammography dataset -> {CLASS_NAME}"
    )
    parser.add_argument("--raw_dir", required=True)
    parser.add_argument("--out_dir", default="datasets")
    parser.add_argument(
        "--sizes", nargs="+", default=["128", "256", "512"],
        type=parse_size,
    )
    args = parser.parse_args()
    main(args.raw_dir, args.out_dir, args.sizes)
